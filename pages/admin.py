import streamlit as st
from utils.db import (get_partidos, upsert_partido, agregar_partido,
                      get_all_colaboradores, get_pronosticos_partido,
                      cargar_lista_blanca, get_lista_blanca, eliminar_de_lista_blanca,
                      lista_blanca_activa, agregar_a_lista_blanca, eliminar_colaborador,
                      es_eliminatoria, eliminar_partido)

def _hora_valida(hora: str) -> bool:
    """Valida formato HH:MM"""
    import re
    return bool(re.match(r"^([01]\d|2[0-3]):[0-5]\d$", hora.strip()))

def render():
    st.markdown("### ⚙️ Panel de Administración")
    st.caption("Solo visible para administradores.")

    tab_res, tab_add, tab_col, tab_lb, tab_det = st.tabs(["📝 Ingresar Resultados", "➕ Agregar Partido", "👥 Colaboradores", "🔐 Lista Blanca", "📊 Detalle Colaborador"])

    # ── Tab 1: Resultados ─────────────────────────────────────────────────────
    with tab_res:
        st.markdown("#### Registrar resultado de un partido")
        partidos = get_partidos()
        pendientes = [p for p in partidos if not p["cerrado"]]
        terminados = [p for p in partidos if  p["cerrado"]]

        if not pendientes:
            st.success("✅ Todos los partidos tienen resultado registrado.")
        else:
            opciones = {
                f"{p['equipo_local']} vs {p['equipo_visita']} ({p['fecha']})": p
                for p in pendientes
            }
            sel = st.selectbox("Partido a cerrar", list(opciones.keys()))
            partido = opciones[sel]
            es_elim = es_eliminatoria(partido["fase"])

            c1, c2 = st.columns(2)
            with c1:
                gl = st.number_input(f"Goles {partido['equipo_local']}", min_value=0, max_value=20, key="admin_gl")
            with c2:
                gv = st.number_input(f"Goles {partido['equipo_visita']}", min_value=0, max_value=20, key="admin_gv")

            # Eliminatorias: quién pasa explícitamente
            equipo_pasa = None
            if es_elim:
                st.markdown("---")
                st.markdown("##### 🏆 ¿Quién avanza? *(obligatorio en eliminatorias)*")
                st.caption("Indica el ganador aunque haya habido prórroga o penales.")
                equipo_pasa = st.radio(
                    "Equipo que pasa:",
                    [partido["equipo_local"], partido["equipo_visita"]],
                    horizontal=True,
                    key="admin_pasa"
                )
                if gl == gv:
                    st.info(f"⏱ Empate {gl}-{gv} al 90 min → pasó **{equipo_pasa}** por prórroga/penales")

            if st.button("Registrar resultado y calcular puntos", type="primary"):
                upsert_partido(partido["id"], gl, gv, equipo_pasa=equipo_pasa)
                extra = f" → pasa {equipo_pasa}" if equipo_pasa else ""
                st.success(f"Resultado: {gl} — {gv}{extra}. Puntos actualizados ✅")
                st.rerun()

        if terminados:
            st.markdown("#### Partidos cerrados")
            for p in terminados:
                prons = get_pronosticos_partido(p["id"])
                exactos   = sum(1 for x in prons if x["puntos"] == 3)
                parciales = sum(1 for x in prons if x["puntos"] == 1)
                pasa_txt  = f" → pasa {p.get('equipo_pasa','')}" if p.get("equipo_pasa") else ""
                st.markdown(f"""
                <div class="card" style="padding:12px 18px;">
                    <b>{p['equipo_local']} {p['goles_local']} — {p['goles_visita']} {p['equipo_visita']}</b>
                    {pasa_txt}
                    &nbsp;·&nbsp; {p['fecha']}
                    &nbsp;·&nbsp; <span style="color:#00B37D">✔ {exactos} exactos</span>
                    &nbsp; <span style="color:#D69E2E">≈ {parciales} parciales</span>
                    &nbsp; <span style="color:#718096">{len(prons)} pronósticos</span>
                </div>
                """, unsafe_allow_html=True)

    # ── Tab 2: Agregar partido ────────────────────────────────────────────────
    with tab_add:
        st.markdown("#### Nuevo partido")

        st.caption("Para eliminatorias usa exactamente: Ronda de 32, Octavos de Final, Cuartos de Final, Semifinal, Tercer Lugar, Final")

        fase   = st.text_input("Fase", placeholder="ej. Ronda de 32, Octavos de Final")
        local  = st.text_input("Equipo local",      placeholder="ej. México")
        visita = st.text_input("Equipo visitante",  placeholder="ej. Argentina")

        c1, c2 = st.columns(2)
        with c1:
            fecha = st.date_input("Fecha del partido")
        with c2:
            hora_str = st.text_input("Hora inicio (hora Mérida)", placeholder="ej. 14:00")

        st.caption("⏰ La hora es importante para el bloqueo automático de pronósticos.")

        if st.button("Agregar partido", type="primary"):
            if not all([fase.strip(), local.strip(), visita.strip()]):
                st.warning("Completa fase, equipo local y equipo visitante.")
            elif hora_str and not _hora_valida(hora_str):
                st.warning("Formato de hora incorrecto. Usa HH:MM, ej. 14:00")
            else:
                hora_final = hora_str.strip() if hora_str.strip() else None
                agregar_partido(fase.strip(), local.strip(), visita.strip(),
                                str(fecha), hora_final)
                st.success(f"✅ {local} vs {visita} agregado — {fecha} {hora_final or 'sin hora'}")
                st.rerun()

        st.markdown("---")
        st.markdown("##### 🗑 Eliminar partido")
        st.caption("Solo puedes eliminar partidos sin resultado registrado.")
        todos = get_partidos()
        eliminables = [p for p in todos if not p["cerrado"] and p["goles_local"] is None]
        if not eliminables:
            st.info("No hay partidos eliminables (todos tienen resultado o están cerrados).")
        else:
            opciones_del = {
                f"{p['fase']} — {p['equipo_local']} vs {p['equipo_visita']} ({p['fecha']})": p
                for p in eliminables
            }
            sel_del = st.selectbox("Selecciona partido a eliminar", list(opciones_del.keys()), key="sel_del")
            partido_del = opciones_del[sel_del]
            if st.button("🗑 Eliminar partido", type="secondary"):
                eliminar_partido(partido_del["id"])
                st.success(f"✅ Partido eliminado correctamente.")
                st.rerun()

    # ── Tab 3: Colaboradores ──────────────────────────────────────────────────
    with tab_col:
        st.markdown("#### Colaboradores registrados")
        cols = get_all_colaboradores()
        st.caption(f"Total registrados: {len([c for c in cols if not c['es_admin']])}")
        for c in cols:
            if c["es_admin"]:
                continue
            c1, c2, c3 = st.columns([3, 2, 1])
            with c1:
                st.markdown(f"**{c['nombre']}**")
            with c2:
                st.markdown(f"<span style='color:#718096;font-size:13px;'>Emp. {c['numero_emp']}</span>", unsafe_allow_html=True)
            with c3:
                if st.button("🗑", key=f"del_col_{c['numero_emp']}", help="Eliminar colaborador"):
                    eliminar_colaborador(c["id"])
                    st.success(f"Colaborador {c['nombre']} eliminado ✅")
                    st.rerun()

    # ── Tab 5: Detalle Colaborador ───────────────────────────────────────────
    with tab_det:
        st.markdown("#### 📊 Detalle de pronósticos por colaborador")
        colaboradores = get_all_colaboradores()
        no_admin = [c for c in colaboradores if not c["es_admin"]]
        if not no_admin:
            st.info("No hay colaboradores registrados aún.")
        else:
            opciones = {f"{c['nombre']} (Emp. {c['numero_emp']})": c for c in no_admin}
            sel = st.selectbox("Selecciona colaborador", list(opciones.keys()))
            colab = opciones[sel]

            partidos = get_partidos()
            cerrados = [p for p in partidos if p["cerrado"] or p["goles_local"] is not None]

            if not cerrados:
                st.info("Aún no hay partidos con resultado.")
            else:
                # Métricas
                total_pts = 0
                exactos = parciales = fallidos = sin_pron = 0
                for p in cerrados:
                    prons = get_pronosticos_partido(p["id"])
                    pron = next((x for x in prons if x["colaborador_id"] == colab["id"]), None)
                    if pron:
                        pts = pron.get("puntos") or 0
                        total_pts += pts
                        if pts == 3: exactos += 1
                        elif pts == 1: parciales += 1
                        else: fallidos += 1
                    else:
                        sin_pron += 1

                c1, c2, c3, c4 = st.columns(4)
                c1.metric("🏆 Puntos", total_pts)
                c2.metric("✅ Exactos", exactos)
                c3.metric("🟡 Parciales", parciales)
                c4.metric("⬜ Sin pronóstico", sin_pron)

                st.markdown("---")

                for p in cerrados:
                    prons = get_pronosticos_partido(p["id"])
                    pron  = next((x for x in prons if x["colaborador_id"] == colab["id"]), None)
                    pts   = pron.get("puntos", 0) if pron else None

                    if pts == 3:
                        emoji, color = "✅", "#00B37D"
                    elif pts == 1:
                        emoji, color = "🟡", "#D69E2E"
                    elif pts == 0 and pron:
                        emoji, color = "❌", "#E03131"
                    else:
                        emoji, color = "⬜", "#A0AEC0"

                    pron_txt = f"{pron['goles_local']}-{pron['goles_visita']}" if pron else "Sin pronóstico"
                    real_txt = f"{p['goles_local']}-{p['goles_visita']}"
                    pts_txt  = f"+{pts} pts" if pron and pts else ("0 pts" if pron else "—")

                    st.markdown(f"""
                    <div style="padding:10px 0; border-bottom:1px solid #EDF2F7;
                                display:flex; align-items:center; gap:12px;">
                        <span style="font-size:1.1rem;">{emoji}</span>
                        <div style="flex:1;">
                            <span style="font-size:13px; font-weight:600;">
                                {p['equipo_local']} vs {p['equipo_visita']}
                            </span>
                            <span style="font-size:12px; color:#718096;"> · {p['fase']} · {p['fecha']}</span>
                        </div>
                        <div style="text-align:center; min-width:80px;">
                            <div style="font-size:11px; color:#718096;">Pronóstico</div>
                            <div style="font-weight:600;">{pron_txt}</div>
                        </div>
                        <div style="text-align:center; min-width:80px;">
                            <div style="font-size:11px; color:#718096;">Real</div>
                            <div style="font-weight:600;">{real_txt}</div>
                        </div>
                        <div style="text-align:right; min-width:60px;
                                    font-weight:700; color:{color};">{pts_txt}</div>
                    </div>
                    """, unsafe_allow_html=True)

    # ── Tab 4: Lista Blanca ───────────────────────────────────────────────────
    with tab_lb:
        st.markdown("#### 🔐 Lista Blanca de Empleados")

        activa = lista_blanca_activa()
        if activa:
            st.success("✅ Control de acceso **activo** — solo números en la lista pueden registrarse.")
        else:
            st.warning("⚠️ Lista blanca **vacía** — cualquier número puede registrarse. Sube tu Excel para activar el control.")

        st.markdown("---")
        st.markdown("##### Agregar empleado individual")
        with st.form("form_agregar_emp"):
            fc1, fc2 = st.columns(2)
            with fc1:
                nuevo_num = st.text_input("Número de empleado", placeholder="Ej. 12345")
            with fc2:
                nuevo_nom = st.text_input("Nombre (opcional)", placeholder="Ej. Juan Pérez")
            submitted = st.form_submit_button("➕ Agregar", type="primary", use_container_width=True)
            if submitted:
                if not nuevo_num.strip():
                    st.warning("Ingresa el número de empleado.")
                else:
                    ok = agregar_a_lista_blanca(nuevo_num.strip(), nuevo_nom.strip())
                    if ok:
                        st.success(f"✅ Empleado {nuevo_num} agregado correctamente.")
                        st.rerun()
                    else:
                        st.warning(f"⚠️ El número {nuevo_num} ya está en la lista.")

        st.markdown("---")
        st.markdown("##### Cargar desde Excel")
        st.caption("El archivo debe tener una columna **numero_emp** y opcionalmente **nombre_ref**.")

        archivo = st.file_uploader("Sube tu Excel (.xlsx)", type=["xlsx"])
        if archivo:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                headers = [str(cell.value).strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

                if "numero_emp" not in headers:
                    st.error("❌ El Excel debe tener una columna llamada 'numero_emp'.")
                else:
                    col_num  = headers.index("numero_emp")
                    col_nom  = headers.index("nombre_ref") if "nombre_ref" in headers else None

                    registros = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        num = str(row[col_num]).strip() if row[col_num] else ""
                        nom = str(row[col_nom]).strip() if col_nom is not None and row[col_nom] else ""
                        if num and num != "None":
                            registros.append({"numero_emp": num, "nombre_ref": nom})

                    st.info(f"📋 {len(registros)} empleados encontrados en el Excel.")

                    if st.button("Cargar lista blanca", type="primary"):
                        insertados = cargar_lista_blanca(registros)
                        st.success(f"✅ {insertados} números cargados correctamente.")
                        st.rerun()
            except Exception as e:
                st.error(f"Error al leer el Excel: {e}")

        st.markdown("---")
        st.markdown("##### Empleados autorizados")
        lista = get_lista_blanca()
        empleados = [l for l in lista if l["numero_emp"] != "0000"]
        st.caption(f"Total: {len(empleados)} empleados")

        for emp in empleados:
            c1, c2, c3 = st.columns([2, 3, 1])
            with c1:
                st.markdown(f"**{emp['numero_emp']}**")
            with c2:
                st.markdown(f"{emp.get('nombre_ref') or '—'}")
            with c3:
                if st.button("🗑", key=f"del_{emp['numero_emp']}"):
                    eliminar_de_lista_blanca(emp["numero_emp"])
                    st.rerun()